from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill


class ExcelFormatter:
    """Basic excel formatting style"""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path

    def load_excel_sheet(self, sheet_name: str):
        try:
            excel_file = load_workbook(self.excel_path)
            sheet = excel_file[sheet_name]
            return sheet
        except Exception:
            raise ImportError(
                "Unable to load excel file from path '%s' with sheet name '%s'"
                % self.excel_path,
                sheet_name,
            )

    def column_width_adjustment(self, sheet_name: str) -> None:
        """Adjusts the width of column to its content."""
        sheet = self.load_excel_sheet(sheet_name)
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        sheet.parent.save(self.excel_path)

    def sheet_formatter(
        self,
        sheet_name: str,
        min_row: int = 1,
        max_row: int | str = "max",
        min_col: int = 1,
        max_col: int | str = "max",
        font: dict | None = None,
        alignment: dict | None = None,
        border: dict | None = None,
        fill: dict | None = None,
    ) -> None:
        """Apply desired format style to sheet cells. Use min_row, max_row, min_col
        and max_col to set the sheet cells that should be formatted."""
        sheet = self.load_excel_sheet(sheet_name)
        if max_row == "max":
            max_row = sheet.max_row
        if max_col == "max":
            max_col = sheet.max_column

        font = font if font else {}
        alignment = alignment if alignment else {}
        border = border if border else {}
        fill = fill if fill else {}

        for row in sheet.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
        ):
            for cell in row:
                cell.alignment = Alignment(**alignment)
                cell.font = Font(**font)
                cell.border = Border(**border)
                cell.fill = PatternFill(**fill)

        sheet.parent.save(self.excel_path)

    def column_type_formatter(
        self, sheet_name: str, min_row: int = 2, column_formatter: dict | None = None
    ) -> None:
        """Format column data types as desired. Use column formatter with keys being
        column letters (eg. 'A' for first column) and value being desired format style.
        For formatting many columns with the same style, use 'other' as a key and
        the value will be provided as a style for all unnamed columns."""
        sheet = self.load_excel_sheet(sheet_name)
        column_formatter = column_formatter if column_formatter else {}

        for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row):
            for cell in row:
                col_letter = cell.column_letter
                if col_letter in column_formatter:
                    cell.number_format = column_formatter[col_letter]
                else:
                    cell.number_format = column_formatter.get("other", "General")
        sheet.parent.save(self.excel_path)
